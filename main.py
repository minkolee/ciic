import openpyxl
import sys

def collect_eid(worksheet) -> list:
    result = set()

    for i in range(start_row, end_row + 1):
        result.add(worksheet.cell(row=i, column=eid_column).value.strip())

    result = list(result)
    result.sort()
    return result


def getEIDData(eid_set, max_row, max_column, worksheet) -> dict:
    result = {}

    for each_eid in eid_set:
        result[each_eid] = {}

    # 对于每一行
    for i in range(2,max_row+1):
        # 对于每一列, 先获取eid，然后组装字典
        eid = worksheet.cell(row=i, column=2).value.strip()

        for j in range(5, max_column+1):
            # 如果序号存在于字典中，就更新，否则新增
            cell_value = worksheet.cell(row=i,column=j).value

            if cell_value:
                if j not in result[eid].keys():
                    print('处理第{}行\tEID：{}\t新增\t新值：{}'.format(i,eid, cell_value))
                    result[eid][j] = cell_value
                else:
                    old_value = result[eid][j]
                    result[eid][j] = cell_value+ result[eid][j]
                    print('处理第{}行\tEID：{}\t叠加\t旧值：{}\t新值：{}'.format(i,eid, old_value,result[eid][j]))

    return result


def writeResult(EIDData, worksheet, fileName='result.xlsx'):

    # 删除第二行开始的数据
    worksheet.delete_rows(2, worksheet.max_row + 1)

    # 从第二行开始写入
    row_start = 2

    for each_eid in EIDData:
        worksheet.cell(row=row_start, column=2).value = each_eid

        for each_column, each_value in EIDData[each_eid].items():
            worksheet.cell(row=row_start, column=each_column).value = each_value

        row_start = row_start + 1

    worksheet.parent.save(fileName)


if __name__ == "__main__":

    filename = 'result.xlsx'

    command_length = len(sys.argv)

    if command_length!=2 and command_length !=3:
        print("参数错误，第一个参数为要处理的文件，第二个参数为要生成的文件名称，可以忽略第二个参数")
        sys.exit(0)

    wb = None
    try:
        wb = openpyxl.open(sys.argv[1])

    except FileNotFoundError:
        print("没有找到文件，请检查文件名是否正确")
        sys.exit(1)

    ws = wb.active

    print("读取表格数据......")
    # 数据开始的行数
    start_row = 2
    print("默认数据从第 {} 行开始处理".format(start_row))

    # 数据结束的行数
    end_row = ws.max_row
    print("表格有效行数为：{} 行".format(end_row))

    # EID所在列数
    eid_column = 2
    print("默认EID所在列数：{}".format(eid_column))

    # 最大列数

    end_column = ws.max_column
    print("表格最大列数为：{}".format(end_column))

    print("上述基础参数错误请联系开发者，开始处理数据......")

    # 生成EID列表
    print("正在生成不重复的EID清单")
    eid_list = collect_eid(ws)
    print("不重复的EID是：{}".format(eid_list))

    print("正在生成不重复的EID与对应汇总数据")
    eid_data = getEIDData(eid_list, end_row,end_column, ws)
    print("汇总数据为：{}".format(eid_data))

    print("开始写入数据")
    if command_length==2:
        print("未输入文件名，默认为result.xlsx")

    if command_length==3:
        filename = sys.argv[2]
        print("文件名为 {}".format(filename))

    try:
        writeResult(eid_data,ws, filename)

        print("文件已经成功写入到：{}".format(filename))

    except:
        print("写入文件发生错误，请检查文件名是否有效或目标文件已经打开")



